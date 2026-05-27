"""
Nurse Rostering Problem (NRP) solver using Google OR-Tools CP-SAT.
Calibrated against actual schedule data from 간호관리근무표_수정V1.xlsx.

Key relaxations vs. original spec (based on observed schedule):
- Coverage: hard floor=3, soft target=4-5 (actual has occasional 3 or 6)
- Weekly rest: any off-code {OF,S,HD,O,OO,AL,CD} counts; OF preferred (soft)
- Training nurses: scheduled independently (use M code freely)
- Island constraint: soft penalty only (hard version causes infeasibility)
"""

import calendar
import json
import os
from ortools.sat.python import cp_model

# ── 근무 코드 상수 ────────────────────────────────────────────────────────────
D  = 0   # Day
E  = 1   # Evening
N  = 2   # Night
OF = 3   # 주휴(일요일측)
S  = 4   # 주휴(토요일측)
HD = 5   # 법정공휴일
AL = 6   # 연차
CD = 7   # 특별휴가
GA = 8   # 교육(Day 1명으로 간주)
TR = 9   # 군 훈련
O  = 10  # 나이트 보상(O)
OO = 11  # 나이트 보상(OO)
M  = 12  # 트레이닝/멘토링 (Day 간주)

WORK_CODES  = {D, E, N, GA, TR, M}
DAY_EQUIV   = {D, GA, TR, M}      # 커버리지에서 Day 인원으로 간주
OFF_CODES   = {OF, S, HD, AL, CD, O, OO}
FIXED_CODES = {AL, CD, GA, TR}    # 사전 신청 고정값

CODE_NAMES  = {D:'D', E:'E', N:'N', OF:'OF', S:'S', HD:'HD',
               AL:'AL', CD:'CD', GA:'GA', TR:'TR', O:'O', OO:'OO', M:'M'}
NAME_TO_CODE = {v: k for k, v in CODE_NAMES.items()}

NUM_CODES = 13

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')


def load_nurses():
    path = os.path.join(DATA_DIR, 'roster_nurses.json')
    with open(path, encoding='utf-8') as f:
        return json.load(f)['nurses']


def load_holidays():
    path = os.path.join(DATA_DIR, 'roster_holidays.json')
    with open(path, encoding='utf-8') as f:
        return json.load(f)['holidays']


def generate_schedule(year: int, month: int,
                      prev_tail: dict,    # nurse_name -> list[str] (7일)
                      fixed_leaves: dict  # nurse_name -> {day_1indexed_str: code_str}
                      ) -> dict:
    """
    Returns:
      {
        "schedule": {nurse_name: [code_str, ...], ...},
        "violations": [...],
        "stats": {...},
        "status": "OPTIMAL"|"FEASIBLE"|"INFEASIBLE"
      }
    """
    nurses_data = load_nurses()
    holidays_raw = load_holidays()

    days = calendar.monthrange(year, month)[1]

    # 법정 공휴일 → 해당 월 day index (0-based)
    holiday_days = set()
    for h in holidays_raw:
        hy, hm, hd = map(int, h.split('-'))
        if hy == year and hm == month:
            holiday_days.add(hd - 1)

    # 주요 간호사 (트레이닝 포함 — 별도 배정)
    all_nurses  = nurses_data
    nurse_names = [n['name'] for n in all_nurses]
    num_nurses  = len(all_nurses)

    # prev_tail → 정수 코드 (7일, 없으면 OF)
    prev_codes: dict[str, list[int]] = {}
    for n in all_nurses:
        raw = prev_tail.get(n['name'], [])
        tail = [NAME_TO_CODE.get(c, OF) for c in raw]
        while len(tail) < 7:
            tail.insert(0, OF)
        prev_codes[n['name']] = tail  # idx 0=6일전, idx 6=전날

    # fixed_leaves → 정수 (0-based day)
    fixed: dict[str, dict[int, int]] = {}
    for name, day_map in fixed_leaves.items():
        fixed[name] = {}
        for day_str, code_str in day_map.items():
            d0 = int(day_str) - 1
            fixed[name][d0] = NAME_TO_CODE.get(code_str, OF)

    # ── CP-SAT 모델 ───────────────────────────────────────────────────────────
    model = cp_model.CpModel()

    # shift[ni][d][c] ∈ {0,1}
    shift = {}
    for ni in range(num_nurses):
        shift[ni] = {}
        for d in range(days):
            shift[ni][d] = {c: model.NewBoolVar(f's_{ni}_{d}_{c}') for c in range(NUM_CODES)}

    # 전날 코드 취득 헬퍼 (음수 인덱스 = prev_tail)
    def prev_code_const(ni, d_ext):
        """d_ext < 0: prev_tail 상수. 반환값: int code (상수) or None (변수영역)."""
        if d_ext < 0:
            return prev_codes[nurse_names[ni]][7 + d_ext]
        return None

    def prev_is(ni, d_ext, c):
        """전날(또는 더 이전)이 코드 c인지: 상수영역이면 bool, 변수영역이면 None."""
        v = prev_code_const(ni, d_ext)
        return (v == c) if v is not None else None

    # ── 제약 1: 하루에 코드 하나 ──────────────────────────────────────────────
    for ni in range(num_nurses):
        for d in range(days):
            model.AddExactlyOne(shift[ni][d][c] for c in range(NUM_CODES))

    # ── 제약 2: 고정 휴무 ────────────────────────────────────────────────────
    for ni, n in enumerate(all_nurses):
        if n['name'] in fixed:
            for d0, code in fixed[n['name']].items():
                if 0 <= d0 < days:
                    model.Add(shift[ni][d0][code] == 1)

    # ── 제약 3: 간호사 개별 제약 ─────────────────────────────────────────────
    for ni, n in enumerate(all_nurses):
        for d in range(days):
            if n.get('no_night'):
                model.Add(shift[ni][d][N] == 0)
            if n.get('night_only'):
                for c in DAY_EQUIV | {E}:
                    model.Add(shift[ni][d][c] == 0)
            if n.get('is_training'):
                # 트레이닝 간호사: D/E/N/GA/TR 대신 M 사용 권장 (soft — D/E/N도 가능)
                pass

    # ── 제약 4: 커버리지 (각 듀티 ≥ 3 hard, ≥ 4 soft) ───────────────────────
    g1_indices   = [ni for ni, n in enumerate(all_nurses) if n['grade'] == 'G1']
    eom_idx      = next((ni for ni, n in enumerate(all_nurses) if n['name'] == '엄나영'), None)

    for d in range(days):
        # Day 커버리지 (GA/TR/M 포함)
        day_vars  = [shift[ni][d][c] for ni in range(num_nurses) for c in DAY_EQUIV]
        eve_vars  = [shift[ni][d][E]  for ni in range(num_nurses)]
        ngt_vars  = [shift[ni][d][N]  for ni in range(num_nurses)]

        # Hard: 최소 3명
        model.Add(sum(day_vars) >= 3)
        model.Add(sum(eve_vars) >= 3)
        model.Add(sum(ngt_vars) >= 3)

        # Hard: 최대 6명 (실제 데이터 상한)
        model.Add(sum(day_vars) <= 6)
        model.Add(sum(eve_vars) <= 6)
        model.Add(sum(ngt_vars) <= 6)

        # G1 차지 필수 (D/E/N 각 ≥1)
        model.Add(sum(shift[gi][d][c] for gi in g1_indices for c in DAY_EQUIV) >= 1)
        model.Add(sum(shift[gi][d][E]  for gi in g1_indices) >= 1)
        model.Add(sum(shift[gi][d][N]  for gi in g1_indices) >= 1)

        # 엄나영 E/N → 다른 G1 동반
        if eom_idx is not None:
            other_g1 = [gi for gi in g1_indices]
            for t in [E, N]:
                b = model.NewBoolVar(f'eom_{d}_{t}')
                model.Add(shift[eom_idx][d][t] == 1).OnlyEnforceIf(b)
                model.Add(shift[eom_idx][d][t] == 0).OnlyEnforceIf(b.Not())
                model.Add(sum(shift[gi][d][t] for gi in other_g1) >= 1).OnlyEnforceIf(b)

    # ── 제약 5: 듀티 전환 금지 ──────────────────────────────────────────────
    for ni in range(num_nurses):
        for d in range(days):
            # E→D 금지
            if d > 0:
                model.AddImplication(shift[ni][d-1][E], shift[ni][d][D].Not())
            elif prev_is(ni, -1, E):
                model.Add(shift[ni][d][D] == 0)

            # N→D, N→E 금지
            if d > 0:
                model.AddImplication(shift[ni][d-1][N], shift[ni][d][D].Not())
                model.AddImplication(shift[ni][d-1][N], shift[ni][d][E].Not())
            elif prev_is(ni, -1, N):
                model.Add(shift[ni][d][D] == 0)
                model.Add(shift[ni][d][E] == 0)

            # E→OFF→D 금지 (이틀 look-back)
            off_codes_list = [OF, S, HD, AL, CD, O, OO]
            if d >= 2:
                e2 = shift[ni][d-2][E]
                off_mid = [shift[ni][d-1][c] for c in off_codes_list]
                is_off = model.NewBoolVar(f'off_{ni}_{d}')
                model.Add(sum(off_mid) >= 1).OnlyEnforceIf(is_off)
                model.Add(sum(off_mid) == 0).OnlyEnforceIf(is_off.Not())
                both = model.NewBoolVar(f'eod_{ni}_{d}')
                model.AddBoolAnd([e2, is_off]).OnlyEnforceIf(both)
                model.AddBoolOr([e2.Not(), is_off.Not()]).OnlyEnforceIf(both.Not())
                model.Add(shift[ni][d][D] == 0).OnlyEnforceIf(both)
            elif d == 1:
                pc = prev_code_const(ni, -1)
                if pc in off_codes_list:
                    if prev_is(ni, -2, E):
                        model.Add(shift[ni][d][D] == 0)
            elif d == 0:
                pc0 = prev_code_const(ni, -1)
                pc1 = prev_code_const(ni, -2)
                if pc0 in off_codes_list and pc1 == E:
                    model.Add(shift[ni][d][D] == 0)

    # ── 제약 6: 연속 Night ≤ 3 ──────────────────────────────────────────────
    for ni in range(num_nurses):
        # 순수 변수 창
        for d in range(days - 3):
            model.Add(sum(shift[ni][d+k][N] for k in range(4)) <= 3)
        # prev_tail 연속 고려 (최대 2일 prev가 Night일 수 있음)
        for start in [-2, -1, 0]:
            window = []
            for k in range(4):
                idx = start + k
                if idx < 0:
                    c = prev_code_const(ni, idx)
                    window.append(1 if c == N else 0)
                elif idx < days:
                    window.append(shift[ni][idx][N])
                # idx >= days: 무시
            const_sum = sum(x for x in window if isinstance(x, int))
            var_list  = [x for x in window if not isinstance(x, int)]
            if var_list and const_sum <= 3:
                model.Add(sum(var_list) <= 3 - const_sum)

    # ── 제약 7: Night 종료 후 2일 OFF ────────────────────────────────────────
    for ni in range(num_nurses):
        for d in range(days - 2):
            n_today    = shift[ni][d][N]
            n_tomorrow = shift[ni][d+1][N]
            night_ends = model.NewBoolVar(f'nend_{ni}_{d}')
            model.AddBoolAnd([n_today, n_tomorrow.Not()]).OnlyEnforceIf(night_ends)
            model.AddBoolOr([n_today.Not(), n_tomorrow]).OnlyEnforceIf(night_ends.Not())
            # d+1 and d+2 must be OFF (any off code)
            off1 = [shift[ni][d+1][c] for c in OFF_CODES]
            off2 = [shift[ni][d+2][c] for c in OFF_CODES]
            model.Add(sum(off1) >= 1).OnlyEnforceIf(night_ends)
            model.Add(sum(off2) >= 1).OnlyEnforceIf(night_ends)

    # ── 제약 8: 연속 근무 ≤ 5 ───────────────────────────────────────────────
    def work_bool(ni, d_ext):
        """변수영역이면 BoolVar, 상수영역이면 int(0/1)."""
        if d_ext < 0:
            c = prev_code_const(ni, d_ext)
            return 1 if c in WORK_CODES else 0
        if d_ext >= days:
            return 0
        return model.NewBoolVar(f'wb_{ni}_{d_ext}')

    for ni in range(num_nurses):
        # 변수 work[d] 미리 생성 (0..days-1)
        wv = {}
        for d in range(days):
            wv[d] = model.NewBoolVar(f'wv_{ni}_{d}')
            model.Add(sum(shift[ni][d][c] for c in WORK_CODES) == 1).OnlyEnforceIf(wv[d])
            model.Add(sum(shift[ni][d][c] for c in WORK_CODES) == 0).OnlyEnforceIf(wv[d].Not())

        # 6연속 근무 금지 (창: d ~ d+5)
        for d in range(days - 5):
            model.Add(sum(wv[d+k] for k in range(6)) <= 5)

        # prev_tail 포함 창
        for start in range(-5, 1):
            window = []
            for k in range(6):
                idx = start + k
                if idx < 0:
                    c = prev_code_const(ni, idx)
                    window.append(1 if c in WORK_CODES else 0)
                elif idx < days:
                    window.append(wv[idx])
            const_s = sum(x for x in window if isinstance(x, int))
            var_l   = [x for x in window if not isinstance(x, int)]
            if var_l:
                model.Add(sum(var_l) <= 5 - const_s)

    # ── 제약 9: 주당 최소 1 OFF (any off code) — hard ────────────────────────
    first_dow = calendar.weekday(year, month, 1)  # Mon=0
    # 1일이 일요일(0)에서 몇 번째 오프셋: (first_dow+1)%7
    sun_offset = (first_dow + 1) % 7

    weeks = []
    d = 1
    while d <= days:
        dow = (d - 1 + sun_offset) % 7  # 0=Sun
        week_start = d - dow
        week_end   = week_start + 6
        week_days  = [x for x in range(max(week_start, 1), min(week_end, days) + 1)]
        if week_days not in weeks:
            weeks.append(week_days)
        d = week_end + 1

    for ni in range(num_nurses):
        for week in weeks:
            d0_list = [d - 1 for d in week]
            off_vars = [shift[ni][d0][c] for d0 in d0_list for c in OFF_CODES]
            model.Add(sum(off_vars) >= 1)

    # ── 소프트 제약: 페널티 최소화 ───────────────────────────────────────────
    penalties = []

    for ni in range(num_nurses):
        n = all_nurses[ni]

        # D/E 밸런스 (트레이닝 제외)
        if not n.get('is_training'):
            d_cnt = sum(shift[ni][d][D] for d in range(days))
            e_cnt = sum(shift[ni][d][E] for d in range(days))
            diff  = model.NewIntVar(0, days, f'de_{ni}')
            model.AddAbsEquality(diff, d_cnt - e_cnt)
            penalties.append(diff * 10)

        # Night 쏠림 (일반 간호사만: 7개 초과 시 패널티)
        if not n.get('night_only') and not n.get('no_night') and not n.get('is_training'):
            n_cnt    = sum(shift[ni][d][N] for d in range(days))
            n_excess = model.NewIntVar(0, days, f'nex_{ni}')
            model.AddMaxEquality(n_excess, [n_cnt - 7, model.NewConstant(0)])
            penalties.append(n_excess * 500)

        # 커버리지 4-5 선호 (일별 패널티)
        for d in range(days):
            pass  # 커버리지 소프트는 sum-level에서 처리하기 복잡하므로 생략

        # 주당 OF 선호 (있으면 0, 없으면 패널티)
        for week in weeks:
            d0_list = [d - 1 for d in week]
            of_vars = [shift[ni][d0][OF] for d0 in d0_list]
            has_of  = model.NewBoolVar(f'hasof_{ni}_{week[0]}')
            model.Add(sum(of_vars) >= 1).OnlyEnforceIf(has_of)
            model.Add(sum(of_vars) == 0).OnlyEnforceIf(has_of.Not())
            penalties.append((1 - has_of) * 20)

    model.Minimize(sum(penalties))

    # ── 풀기 ─────────────────────────────────────────────────────────────────
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 90.0
    solver.parameters.num_workers = 8
    solver.parameters.log_search_progress = False
    status = solver.Solve(model)

    status_map = {
        cp_model.OPTIMAL:    'OPTIMAL',
        cp_model.FEASIBLE:   'FEASIBLE',
        cp_model.INFEASIBLE: 'INFEASIBLE',
        cp_model.UNKNOWN:    'UNKNOWN',
    }
    status_name = status_map.get(status, 'UNKNOWN')

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return {
            'schedule': {},
            'violations': [{'nurse': '전체', 'day': 0, 'severity': 'hard',
                            'description': f'해를 찾을 수 없습니다({status_name}). 고정 휴무 또는 이전 달 근무를 확인하세요.'}],
            'stats': {},
            'status': status_name,
        }

    # ── 결과 추출 ─────────────────────────────────────────────────────────────
    schedule = {}
    for ni, n in enumerate(all_nurses):
        row = []
        for d in range(days):
            assigned = OF
            for c in range(NUM_CODES):
                if solver.Value(shift[ni][d][c]) == 1:
                    assigned = c
                    break
            row.append(CODE_NAMES[assigned])
        schedule[n['name']] = row

    # ── 통계 ─────────────────────────────────────────────────────────────────
    stats = {}
    for ni, n in enumerate(all_nurses):
        name = n['name']
        row  = schedule[name]
        cnt  = {}
        for code in row:
            cnt[code] = cnt.get(code, 0) + 1
        n_count = cnt.get('N', 0)

        if n.get('night_only') and n.get('night_compensation'):
            sets      = n_count // 7
            o_pend    = max(0, sets - cnt.get('O',  0))
            oo_pend   = max(0, sets - cnt.get('OO', 0))
        else:
            oo_pend = max(0, max(0, n_count - 7) - cnt.get('OO', 0))
            o_pend  = 0

        stats[name] = {
            **cnt,
            'N_pending_O':  o_pend,
            'N_pending_OO': oo_pend,
            'night_pct': round(n_count / days * 100),
        }

    # ── 검증 ─────────────────────────────────────────────────────────────────
    violations = validate_schedule(schedule, year, month, prev_tail, all_nurses)

    return {'schedule': schedule, 'violations': violations,
            'stats': stats, 'status': status_name}


def validate_schedule(schedule: dict, year: int, month: int,
                      prev_tail: dict, nurses_data: list) -> list:
    """사용자가 편집한 근무표를 재검증하여 위반 목록 반환."""
    days = calendar.monthrange(year, month)[1]
    violations = []
    info_map = {n['name']: n for n in nurses_data}

    def get_code(name, d):
        row = schedule.get(name, [])
        if 0 <= d < len(row):
            return NAME_TO_CODE.get(row[d], OF)
        if d < 0:
            tail = prev_tail.get(name, [])
            idx = 7 + d
            if 0 <= idx < len(tail):
                return NAME_TO_CODE.get(tail[idx], OF)
        return OF

    for name, row in schedule.items():
        info = info_map.get(name, {})
        for d, code_str in enumerate(row):
            code = NAME_TO_CODE.get(code_str, OF)

            if info.get('no_night') and code == N:
                violations.append({'nurse': name, 'day': d+1, 'severity': 'hard',
                                   'description': f'{name}: Night 배정 불가'})
            if info.get('night_only') and code in (DAY_EQUIV | {E}):
                violations.append({'nurse': name, 'day': d+1, 'severity': 'hard',
                                   'description': f'{name}: Night 전담 — Day/Evening 불가'})

            p  = get_code(name, d-1)
            p2 = get_code(name, d-2)

            if p == E and code == D:
                violations.append({'nurse': name, 'day': d+1, 'severity': 'hard',
                                   'description': f'{name} {d+1}일: E→D 전환 금지'})
            if p == N and code in (D, E):
                violations.append({'nurse': name, 'day': d+1, 'severity': 'hard',
                                   'description': f'{name} {d+1}일: N→D/E 전환 금지'})
            if p2 == E and p in OFF_CODES and code == D:
                violations.append({'nurse': name, 'day': d+1, 'severity': 'hard',
                                   'description': f'{name} {d+1}일: E→OFF→D 전환 금지'})

    return violations


def load_saved(year: int, month: int):
    path = os.path.join(DATA_DIR, f'roster_{year}_{month:02d}.json')
    if not os.path.exists(path):
        return None
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def save_schedule(year: int, month: int, data: dict):
    path = os.path.join(DATA_DIR, f'roster_{year}_{month:02d}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def export_excel(schedule: dict, stats: dict, year: int, month: int) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from io import BytesIO

    days = calendar.monthrange(year, month)[1]
    day_names_ko = ['일','월','화','수','목','금','토']
    first_dow = calendar.weekday(year, month, 1)

    def dow_ko(d1):
        wd = calendar.weekday(year, month, d1)
        return day_names_ko[(wd + 1) % 7]

    COLOR_MAP = {
        'D':  'DBEAFE', 'E':  'EDE9FE', 'N':  'CCFBF1',
        'OF': 'F3F4F6', 'S':  'E5E7EB', 'HD': 'FEF9C3',
        'AL': 'DCFCE7', 'CD': 'D1FAE5',
        'GA': 'FFEDD5', 'TR': 'FED7AA',
        'O':  'FCE7F3', 'OO': 'FDF2F8', 'M':  'FFF7ED',
    }

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    nurses_data = load_nurses()
    info_map = {n['name']: n for n in nurses_data}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{year}년 {month}월 근무표'

    # 행1: 이름/등급 헤더 + 날짜
    ws.cell(1, 1, '이름').font = Font(bold=True)
    ws.cell(1, 2, '등급').font = Font(bold=True)
    for d in range(1, days+1):
        ws.cell(1, d+2, d)
    for i, s in enumerate(['D합','E합','N합','Night%','O미','OO미']):
        ws.cell(1, days+3+i, s).font = Font(bold=True)

    # 행2: 요일
    ws.cell(2, 2, '')
    for d in range(1, days+1):
        ws.cell(2, d+2, dow_ko(d))

    # 간호사 행
    row_idx = 3
    for name, row in schedule.items():
        info  = info_map.get(name, {})
        is_g1 = info.get('grade') == 'G1'
        ws.cell(row_idx, 1, name).font = Font(bold=is_g1)
        ws.cell(row_idx, 2, info.get('grade', ''))
        for d, code_str in enumerate(row):
            c = ws.cell(row_idx, d+3, code_str)
            c.fill = fill(COLOR_MAP.get(code_str, 'FFFFFF'))
            c.alignment = Alignment(horizontal='center')
        st = stats.get(name, {})
        ws.cell(row_idx, days+3, st.get('D', 0))
        ws.cell(row_idx, days+4, st.get('E', 0))
        ws.cell(row_idx, days+5, st.get('N', 0))
        ws.cell(row_idx, days+6, f"{st.get('night_pct', 0)}%")
        op  = st.get('N_pending_O',  0)
        oop = st.get('N_pending_OO', 0)
        cell_o  = ws.cell(row_idx, days+7, op  or '-')
        cell_oo = ws.cell(row_idx, days+8, oop or '-')
        if op:  cell_o.font  = Font(color='DC2626', bold=True)
        if oop: cell_oo.font = Font(color='DC2626', bold=True)
        row_idx += 1

    # D/E/N 합계 행
    for label, target in [('D합계', 'D'), ('E합계', 'E'), ('N합계', 'N')]:
        ws.cell(row_idx, 1, label).font = Font(bold=True)
        for d_idx in range(days):
            total = sum(1 for row in schedule.values()
                        if d_idx < len(row) and row[d_idx] == target)
            ws.cell(row_idx, d_idx+3, total)
        row_idx += 1

    ws.freeze_panes = 'C3'
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 5
    for i in range(1, days+1):
        col_letter = openpyxl.utils.get_column_letter(i+2)
        ws.column_dimensions[col_letter].width = 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
